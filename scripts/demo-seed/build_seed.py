"""
Bouwt de demo-organisatie voor Koninklijke Stade Leuven Tennis in Kaspio.

Twee bronnen:
  - "2026 budget en overzicht in-uit opvolging - Kaspio.xlsx"
      BudgetOverzicht  -> comites (pot_groups) + posten (pots) + budget 2026
      comite-bladen    -> bankregels die al per post gesorteerd zijn
  - "Cashflow sinds 2004.xlsx", blad 2026
      -> de aanzuiveringen tussen eigen rekeningen, die in de comite-bladen
         bewust ontbreken omdat ze geen inkomst of uitgave zijn

De koppeling transactie -> potje komt dus uit het bestand zelf, precies zoals de
potje-kolom van de import op feat/blok-0-import-rekeningen bedoeld is.
"""

import openpyxl, re, json, datetime, collections, uuid, os

BUD = "/Users/tuylss/Downloads/2026 budget en overzicht in-uit opvolging - Kaspio.xlsx"
CF = "/Users/tuylss/Downloads/Cashflow sinds 2004.xlsx"
OUT = os.path.dirname(os.path.abspath(__file__))

ORG_NAME = "Koninklijke Stade Leuven Tennis"
OWNER = "8ce38e95-6a4f-4d5c-a659-8ef96fb3c023"  # stormtuyls@icloud.com

REF = re.compile(r"^\d{4}-\d{4,5}$")
YEAR = 2026

# Blad -> comité(s) waarvan de posten op dat blad mogen voorkomen. Het is
# meestal één op één, maar niet altijd: het blad "Gtennis" bevat zowel de
# rolstoel- als de VE-posten, en in BudgetOverzicht staan die onder twee
# verschillende comités. Zonder de tweede naam vallen die rijen terug op de
# vorige post en belanden ze op het verkeerde potje.
SHEET2COMITE = {
    "Dagelijks Bestuur": ["Dagelijks bestuur"], "Zomertennis": ["Zomertennis"],
    "Subsidies": ["Subsidies"], "Gtennis": ["Rolstoeltennis", "VE-tennis"],
    "Bar": ["Bar"], "Tennisschool": ["Tennisschool"],
    "Wintertennis": ["Wintertennis"], "Competitie": ["Competitie"],
    "Infrastructuur": ["Infrastructuur"], "Sponsoring": ["Sponsoring"],
    "P.R.": ["Public Relations"], "Jeugd & feest": ["Jeugd- en feestcomité"],
    "Financiën": ["Financiën"],
}

# Bladrubrieken die bij een andere budgetpost horen dan waar ze onder staan.
#
# "Loon Barman" staat op het blad onder het roll-upkopje TOTAAL LOONKOST, en
# belandde daardoor in de post "TOTAAL LOONKOST Gert Vanderlinden". Dat is het
# loon van Sven Swinnen, niet van Gert, en er bestaat een eigen budgetpost voor
# die tot nu leeg bleef. Zo'n uitzondering hoort expliciet te staan in plaats
# van in een heuristiek te verdwijnen.
#
# Gevolg: de post van Gert komt niet meer uit op zijn Stand 2026, want dat
# cijfer telt de barman mee. Dat is de prijs voor een juiste naam boven het
# juiste bedrag.
RUBRIEK_NAAR_POST = {
    ("Dagelijks bestuur", "LOON BARMAN"): "TOTALE LOONKOST VAST BARPERSONEEL",
}

# Eén kleur per comité, zodat het overzicht per tak leesbaar blijft.
GROUP_COLOR = {
    "Dagelijks bestuur": "#1D9E75", "Zomertennis": "#F59E0B",
    "Wintertennis": "#3B82F6", "Tennisschool": "#8B5CF6", "Bar": "#EF4444",
    "Infrastructuur": "#0EA5E9", "Competitie": "#EC4899",
    "Public Relations": "#14B8A6", "Sponsoring": "#A3E635",
    "Subsidies": "#22C55E", "Jeugd- en feestcomité": "#FB923C",
    "Rolstoeltennis": "#6366F1", "VE-tennis": "#D946EF", "Financiën": "#64748B",
}

GROUP_ORDER = ["Dagelijks bestuur", "Financiën", "Zomertennis", "Wintertennis",
               "Tennisschool", "Competitie", "Bar", "Infrastructuur",
               "Sponsoring", "Subsidies", "Public Relations",
               "Jeugd- en feestcomité", "Rolstoeltennis", "VE-tennis"]


def norm(s):
    return re.sub(r"\s+", " ", str(s)).strip().upper()


def q(s):
    """SQL string literal, of NULL."""
    if s is None:
        return "null"
    s = re.sub(r"\s+", " ", str(s)).strip()
    if not s:
        return "null"
    return "'" + s.replace("'", "''") + "'"


def acc(s):
    return re.sub(r"\s+", "", str(s or "")).upper()


# =============================================================================
# 1. Budgetposten
# =============================================================================
wb = openpyxl.load_workbook(BUD, read_only=True, data_only=True)
ws = wb["BudgetOverzicht"]
comite = None
posts = []          # [{comite, post, budget2026, ...}]
seen_post = set()
for r in ws.iter_rows(min_row=3, values_only=True):
    c = (r[0] or "").strip() if isinstance(r[0], str) else ""
    p = (r[2] or "").strip() if isinstance(r[2], str) else ""
    if c:
        comite = c
    if not p:
        continue
    key = (comite, norm(p))
    if key in seen_post:
        continue
    seen_post.add(key)
    posts.append(dict(comite=comite, post=p, key=norm(p),
                      stand=r[9], budget=r[10], prog=r[11],
                      w2024=r[7], w2025=r[8]))

# Vaste ids uit de naam, zodat opnieuw genereren dezelfde uuids oplevert en een
# hergedraaide seed te vergelijken blijft met wat er al in de databank staat.
NS = uuid.UUID("6f1d5a3e-9c2b-4a77-9f0e-3b1c2d4e5f60")


def det(*parts):
    return str(uuid.uuid5(NS, "|".join(str(p) for p in parts)))


pot_id = {(p["comite"], p["key"]): det("pot", p["comite"], p["key"]) for p in posts}
group_id = {g: det("group", g) for g in GROUP_ORDER}
known_posts = collections.defaultdict(set)
for p in posts:
    known_posts[p["comite"]].add(p["key"])

# =============================================================================
# 2. Bankregels uit de comité-bladen, met hun post
# =============================================================================
# De bladen zijn duidelijk: boven elke reeks staat een kopje, en die reeks hoort
# bij dat kopje. Alle 2.326 bankregels staan onder een kopje, zonder uitzondering.
# Dat is de categorisering van de club zelf, en die gooien we niet weg.
#
# Wat wél schuurt is de naamgeving. Van de 97 kopjes met regels zijn er 57 op een
# budgetpost te herleiden; de rest zijn tussenkopjes ("LOON SECRETARIS",
# "BANCONTACT", "COLRUYT") die onder een post vallen of er los naast staan.
#
# Vandaar deze aanpak per blad:
#   1. Kopjes die een budgetpost zijn, beginnen een post.
#   2. De tussenkopjes daarna zijn kandidaat-kinderen. We nemen precies zoveel
#      van hun regels mee als nodig om de post exact op "Stand 2026" te laten
#      uitkomen. Klopt het al zonder, dan nemen we er geen.
#   3. Wat overblijft krijgt een eigen potje met de naam van het blad. Dat is
#      beter dan doorlekken op de vorige post: zo belandde het Mollie-blok van
#      130.000 euro vroeger op VTV NIET-LEDENAANSLUITINGEN.
#
# Afsluitingsregels (geen referte, geen rekening: "LIDGELDEN ZOMER") doen mee in
# stap 2, want ze zitten in hun Stand 2026. Ze zijn geen bankverrichting, dus ze
# worden geboekt als interne overboeking: één been op de post, een tegenboeking
# op de hoofdpot. Zo klopt de post én blijft het totaal gelijk aan de rekening.

def as_num(x):
    return x if isinstance(x, (int, float)) else None


def cent(x):
    return int(round(float(x) * 100))


recs = []          # bankregels met een post
afsl = []          # afsluitingsregels met een post -> worden overboekingen
losse = []         # niets van te maken -> hoofdpot, onbeslist
extra_pots = []    # (comite, naam) voor kopjes die geen budgetpost zijn
n_exact = 0
n_post = 0

for sh, coms in SHEET2COMITE.items():
    hoofdcom = coms[0]
    by_name, by_stand = {}, collections.defaultdict(list)
    for p in posts:
        if p["comite"] not in coms:
            continue
        by_name.setdefault(p["key"], p)
        s = as_num(p["stand"])
        if s is not None and round(s, 2) != 0:
            by_stand[round(s, 2)].append(p)
    by_stand = {k: v[0] for k, v in by_stand.items() if len(v) == 1}

    # Blad in secties knippen: [naam, subtotaal, post-of-None, regels]
    secties = []
    for r in wb[sh].iter_rows(values_only=True):
        a = r[0]
        if isinstance(a, str) and REF.match(a.strip()):
            if as_num(r[3]) and r[3] and secties:
                secties[-1][3].append(dict(
                    soort="bank", ref=a.strip(), date=r[1], amount=r[3], own=r[5],
                    kind=r[6], cp_acc=r[7], cp_name=r[8], memo=r[9], descr=r[10]))
        elif isinstance(a, str) and a.strip() and not isinstance(r[1], datetime.datetime):
            n, tot, bedrag = norm(a), as_num(r[4]), as_num(r[3])
            # Een kopje zet zijn subtotaal in kolom E. Staat er in plaats
            # daarvan een bedrag in kolom D (dezelfde kolom als bankregels),
            # dan is dit geen kopje maar een handgeschreven regel: "KOCH Anja
            # -20" onder GRATIS LIDGELDEN. Die werden hiervoor als kopje
            # gelezen en dus weggegooid, 120 stuks.
            if tot is None and bedrag:
                if secties:
                    secties[-1][3].append(dict(
                        soort="handmatig", ref=None, date=None, amount=bedrag,
                        own=None, kind="handmatig", cp_acc=None,
                        cp_name=a.strip(), memo=a.strip(), descr=None))
                continue
            hit = None
            for c in coms:
                doel = RUBRIEK_NAAR_POST.get((c, n))
                if doel:
                    hit = by_name.get(doel)
                    break
            if hit is None:
                hit = by_name.get(n)
            if hit is None and tot is not None:
                hit = by_stand.get(round(tot, 2))
            secties.append([n, tot, hit, []])
        elif a in (None, "") and isinstance(r[1], str) and r[1].strip() \
                and isinstance(r[3], (int, float)) and r[3] and secties:
            # Maandregels: kolom A leeg, de maand in kolom B, het bedrag in
            # kolom D. Zo houdt de club posten bij die per maand worden
            # afgerekend (de RvB-onkosten, het barpersoneel). Mijn parser keek
            # alleen naar kolom A en zag deze 41 rijen dus niet.
            note = r[4] if isinstance(r[4], str) else None
            label = r[1].strip() + (f" ({note.strip()})" if note else "")
            secties[-1][3].append(dict(
                soort="handmatig", ref=None, date=None, amount=r[3], own=None,
                kind="handmatig", cp_acc=None, cp_name=secties[-1][0].title(),
                memo=label, descr=None))
        elif a in (None, "") and isinstance(r[2], datetime.datetime) \
                and isinstance(r[3], (int, float)) and r[3] and secties:
            lbl = r[4] if isinstance(r[4], str) else None
            secties[-1][3].append(dict(
                soort="afsluiting", ref=None, date=r[2], amount=r[3], own=None,
                kind="afsluiting", cp_acc=None, cp_name=lbl, memo=lbl, descr=None))

    # Onder "AFSCHRIJVINGEN" staat in kolom D het afschrijvingsjaar, niet een
    # bedrag: 2026, 2025, 2024. Blind optellen gaf daar 117.177 euro aan onzin.
    # Een sectie waarvan vrijwel elke handmatige waarde een rond getal tussen
    # 2000 en 2100 is, gaat dus over jaren.
    for sec in secties:
        hand = [x for x in sec[3] if x["soort"] == "handmatig"]
        if not hand:
            continue
        jaren = sum(1 for x in hand
                    if float(x["amount"]).is_integer() and 2000 <= x["amount"] <= 2100)
        if jaren / len(hand) >= 0.8:
            sec[3] = [x for x in sec[3] if x["soort"] != "handmatig"]

    # Roll-upkopjes: een kopje met een eigen subtotaal maar zonder eigen regels,
    # waarvan de kopjes erna optellen tot dat bedrag. Dat is het niveau tussen
    # comité en post dat de bladen wel hebben en Kaspio niet: "TOTAALKOST
    # ONDERHOUD", "INVESTERINGEN TOTAAL", "AANKOOP VOEDING".
    #
    # Minstens twee kinderen, en het bedrag mag geen jaartal zijn. Zonder die
    # twee eisen ziet hij in het afschrijvingsblok van Financiën overal ketens
    # die er niet zijn (LUCHTHAL -> LUCHTHAL -> LUCHTHAL).
    i = 0
    while i < len(secties):
        naam, tot, post, rijen = secties[i]
        if post is None:
            # Los kopje zonder post: eigen potje.
            bank = [x for x in rijen if x["soort"] == "bank"]
            if bank:
                sleutel = (hoofdcom, naam)
                if sleutel not in dict.fromkeys(extra_pots):
                    extra_pots.append(sleutel)
                for x in bank:
                    x.update(comite=hoofdcom, post=naam, eigen=True)
                    recs.append(x)
            i += 1
            continue

        n_post += 1
        # De kandidaat-kinderen: alle volgende secties tot de eerstvolgende post.
        # Kandidaat-kinderen: de volgende secties tot de eerstvolgende post.
        # Een sectie met een EIGEN subtotaal in kolom E is echter geen
        # tussenkopje maar een volwaardige rubriek: MASTERPLAN, MINDERVALIDEN,
        # BAR & TERRAS onder INVESTERINGEN, of Loon Barman onder TOTAAL
        # LOONKOST. Die krijgen hun eigen potje, want anders zoekt de club naar
        # een naam die niet bestaat en ziet ze alleen de optelsom.
        #
        # Tussenkopjes zonder eigen bedrag (de maandregels, de personen onder
        # GRATIS LIDGELDEN) rollen wel gewoon op in hun post.
        j = i + 1
        kinderen = []
        while j < len(secties) and secties[j][2] is None:
            if secties[j][1] is not None:
                break
            kinderen.append(j)
            j += 1

        stroom = list(rijen) + [x for k in kinderen for x in secties[k][3]]
        doel = as_num(post["stand"])
        neem = None
        if doel is not None:
            loop = 0
            for t, x in enumerate(stroom, 1):
                loop += cent(x["amount"])
                if abs(loop - cent(round(doel, 2))) < 2:
                    neem = t
                    n_exact += 1
                    break
        if doel is None:
            # Geen subtotaal om tegen af te vinken, dus is er ook niets om
            # kinderen mee te verantwoorden. Alleen de eigen regels nemen.
            # Zonder deze regel slokte een post zonder Stand alles op wat
            # erachter stond: "Loon Barman" pakte ook Loon Secretaris, Soc Secr
            # en Maaltijdcheques mee, 37 regels in plaats van 8.
            neem = len(rijen)
        elif neem is None:
            # Geen enkele combinatie komt op het subtotaal uit. Dan zijn die
            # tussenkopjes toch van deze post: ze staan eronder en er is geen
            # andere post die ze kan opeisen. Alles meenemen dus, en het
            # verschil met Stand 2026 laten zien in plaats van het geld op een
            # potje met een vreemde naam te parkeren. "KOSTEN TRAINERS LENTE"
            # dat 50 euro afwijkt is bruikbaar; datzelfde bedrag onder
            # "CORRECTIE NAAR WAARDEN ICLUB" met een lege trainerspost niet.
            neem = len(stroom)

        for x in stroom[:neem]:
            x.update(comite=post["comite"], post=post["key"], eigen=False)
            (afsl if x["soort"] in ("afsluiting", "handmatig") else recs).append(x)

        # Wat de post niet nodig had, terug naar de kindsecties als eigen potje.
        rest = {id(x) for x in stroom[neem:]}
        for k in kinderen:
            bank = [x for x in secties[k][3] if x["soort"] == "bank" and id(x) in rest]
            if not bank:
                continue
            sleutel = (hoofdcom, secties[k][0])
            if sleutel not in dict.fromkeys(extra_pots):
                extra_pots.append(sleutel)
            for x in bank:
                x.update(comite=hoofdcom, post=secties[k][0], eigen=True)
                recs.append(x)
        i = j

wb.close()

extra_pots = list(dict.fromkeys(extra_pots))
# Deze potjes staan niet in BudgetOverzicht, dus ze krijgen geen budget en geen
# prognose. Ze bestaan omdat de club ze op haar blad als aparte reeks bijhoudt,
# en dat is precies het detail dat anders verloren gaat.
for k in extra_pots:
    pot_id.setdefault(k, det("pot", *k))
print(f"posten die exact op Stand 2026 uitkomen: {n_exact}")
print(f"bankregels toegewezen: {len(recs)}   afsluitingsregels: {len(afsl)}")
print(f"extra potjes uit de bladen: {len(extra_pots)}")


def in_year(d):
    return isinstance(d, datetime.datetime) and d.year == YEAR


# =============================================================================
# 3. Cashflow 2026: de rijen die de comité-bladen niet hebben (aanzuiveringen)
# =============================================================================
def key5(own, date, amt, cp, memo):
    return (acc(own), str(date)[:10], round(float(amt), 2), acc(cp),
            re.sub(r"\s+", " ", str(memo or "")).strip().upper()[:30])

# Let op: alle comité-regels, ook die naar de hoofdpot verhuisd zijn. Anders
# komen die verderop een tweede keer binnen als "cashflow-only".
have = collections.Counter(key5(r["own"], r["date"], r["amount"], r["cp_acc"], r["memo"])
                           for r in (recs + losse) if r["ref"])

wb2 = openpyxl.load_workbook(CF, read_only=True, data_only=True)
extra = []
for r in wb2["2026"].iter_rows(values_only=True):
    if not (isinstance(r[0], str) and REF.match(r[0].strip())):
        continue
    if not isinstance(r[3], (int, float)) or not r[3] or not in_year(r[1]):
        continue
    k = key5(r[5], r[1], r[3], r[7], r[9])
    if have[k] > 0:
        have[k] -= 1
        continue
    extra.append(dict(comite=None, post=None, ref=r[0].strip(), date=r[1],
                      amount=r[3], own=r[5], kind=r[6], cp_acc=r[7],
                      cp_name=r[8], memo=r[9], descr=r[10], source="cashflow"))
wb2.close()

# =============================================================================
# 4. Interne overboekingen koppelen (port van detectInternalTransfers)
# =============================================================================
OWN_ACCOUNTS = {acc(a) for a in [
    "BE72230034114516", "BE02230034437040", "BE31001790004755",
    "BE97230034408849", "BE39230704617619", "BE13035735570239"]}


def day_gap(a, b):
    return abs((a - b).days)


def detect_transfers(rows, window=3):
    partner = [None] * len(rows)
    usable = []
    for r in rows:
        own, other = acc(r["own"]), acc(r["cp_acc"])
        if not own or not other or own == other or other not in OWN_ACCOUNTS:
            usable.append(None)
        else:
            usable.append((own, other, r["amount"], r["date"]))
    for i, a in enumerate(usable):
        if not a or partner[i] is not None:
            continue
        best, best_gap = -1, 10 ** 9
        for j in range(i + 1, len(usable)):
            b = usable[j]
            if not b or partner[j] is not None:
                continue
            if (a[2] > 0) == (b[2] > 0):
                continue
            if round(abs(a[2]), 2) != round(abs(b[2]), 2):
                continue
            if a[0] != b[1] or a[1] != b[0]:
                continue
            g = day_gap(a[3], b[3])
            if g > window or g >= best_gap:
                continue
            best, best_gap = j, g
        if best >= 0:
            partner[i] = best
            partner[best] = i
    return partner


partner = detect_transfers(extra)
tgroup = [None] * len(extra)
npair = 0
for i, j in enumerate(partner):
    if j is not None and tgroup[i] is None:
        g = det("transfer", extra[i]["ref"], str(extra[i]["date"])[:10], extra[i]["amount"])
        tgroup[i] = tgroup[j] = g
        npair += 1

# =============================================================================
# 5. Leden
# =============================================================================
chairs = json.load(open(os.path.join(OUT, "chairs.json")))
people = {}
for com, name in chairs.items():
    people.setdefault(name, []).append(com)


def mail(name):
    parts = re.sub(r"[^a-zA-Z ]", "", name).lower().split()
    return ".".join(parts) + "@demo.kaspio.be"


user_id = {n: det("user", n) for n in people}

# =============================================================================
# 6. SQL genereren
# =============================================================================
ORG = det("org", ORG_NAME)
# De tegenboekingen van de afsluitingsregels moeten ergens staan. Niet op de
# hoofdpot: die betekent "geld dat nog een bestemming moet krijgen", en dit geld
# is juist al bestemd, het is alleen intern verschoven. Zet je het er toch op,
# dan leest het dashboard 115.830 euro als verdeelbaar terwijl er niets te
# verdelen valt. Bewust zonder groep, want het hoort bij geen enkel comité.
HERVERDEEL = det("pot", "__herverdeling__")
# Het beginsaldo hoort ook niet in de hoofdpot. "Nog te verdelen" betekent geld
# dat nog een bestemming zoekt; de kaspositie van 31 december is dat niet, dat
# is de reserve waarmee de club het jaar begint. In de hoofdpot leest het als
# 73.000 euro vrij besteedbaar, en dat is het niet.
RESERVE = "a1e7c3d4-0000-5000-9000-000000000001"
sql = []
A = sql.append

A(f"""-- Demo-org {ORG_NAME}
-- Opnieuw draaibaar: gooit een vorige versie van deze demo-org eerst weg.
do $$
declare v_org uuid;
begin
  for v_org in select id from public.organisations
                where name = {q(ORG_NAME)} and owner_id = '{OWNER}'
  loop
    delete from public.allocations where organisation_id = v_org;
    delete from public.transactions where organisation_id = v_org;
    delete from public.recurring_plans where organisation_id = v_org;
    delete from public.distribution_shares where organisation_id = v_org;
    delete from public.memberships where organisation_id = v_org;
    delete from public.audit_log where organisation_id = v_org;
    delete from public.pots where organisation_id = v_org and not is_hoofdpot;
    delete from public.pot_groups where organisation_id = v_org;
    delete from public.pots where organisation_id = v_org;
    delete from public.subscriptions where organisation_id = v_org;
    delete from public.organisations where id = v_org;
  end loop;
end $$;

insert into public.organisations (id, name, owner_id)
values ('{ORG}', {q(ORG_NAME)}, '{OWNER}');

-- Team-tier vóór de potjes: de gratis limiet staat op 3 potjes en 2 leden.
update public.subscriptions
   set tier = 'team', status = 'active',
       current_period_end = now() + interval '1 year'
 where organisation_id = '{ORG}';

insert into public.memberships (organisation_id, user_id, role)
values ('{ORG}', '{OWNER}', 'admin');
""")

# --- pot_groups
A("insert into public.pot_groups (id, organisation_id, name, sort_order) values")
A(",\n".join(
    f"  ('{group_id[g]}', '{ORG}', {q(g)}, {i})"
    for i, g in enumerate(GROUP_ORDER)) + ";\n")

# --- pots
rows = []
for p in posts:
    pid = pot_id[(p["comite"], p["key"])]
    b = p["budget"]
    if b is None or not isinstance(b, (int, float)) or round(b, 2) == 0:
        target, kind = "null", "saving"
    elif b < 0:
        target, kind = f"{abs(round(b,2)):.2f}", "budget"
    else:
        target, kind = f"{round(b,2):.2f}", "saving"
    # Prognose 2026 uit hun eigen blad, met dezelfde tekenregel als het budget.
    pr = p["prog"]
    if isinstance(pr, (int, float)) and round(pr, 2) != 0:
        forecast = f"{abs(round(pr,2)):.2f}" if kind == "budget" else f"{round(pr,2):.2f}"
    else:
        forecast = "null"
    desc = f"{p['comite']} · werkelijk 2025: {p['w2025']:,.2f} EUR" \
        if isinstance(p["w2025"], (int, float)) else p["comite"]
    rows.append(f"  ('{pid}', '{ORG}', {q(p['post'][:80])}, "
                f"'{GROUP_COLOR[p['comite']]}', {target}, {forecast}, '{kind}', "
                f"'{group_id[p['comite']]}', {q(desc)})")
rows.append(f"  ('{RESERVE}', '{ORG}', 'Overgedragen uit 2025', "
            f"'#0F766E', null, null, 'saving', null, "
            f"{q('Kaspositie op 31 december 2025. Dit is de reserve waarmee de '
                 'club het jaar begon, geen geld dat nog een bestemming zoekt.')})")

rows.append(f"  ('{HERVERDEEL}', '{ORG}', 'Herverdeling tussen posten', "
            f"'#94A3B8', null, null, 'saving', null, "
            f"{q('Tegenboekingen van de afsluitingsregels uit het rekenblad. '
                 'Samen met het been op de post is elke herverdeling netto nul, '
                 'zodat het totaal gelijk blijft aan de rekening.')})")

for com, naam in extra_pots:
    rows.append(f"  ('{pot_id[(com, naam)]}', '{ORG}', {q(naam[:80])}, "
                f"'{GROUP_COLOR[com]}', null, null, 'saving', "
                f"'{group_id[com]}', {q(com + ' · eigen rubriek uit het comitéblad')})")

A("insert into public.pots (id, organisation_id, name, color, target_amount, "
  "forecast_amount, target_kind, group_id, description) values")
A(",\n".join(rows) + ";\n")

# --- leden
A("-- Comitévoorzitters als pot_owner. Demo-accounts, wachtwoord 'kaspio-demo'.")
for n, uid in user_id.items():
    A(f"""insert into auth.users (id, instance_id, aud, role, email,
    encrypted_password, email_confirmed_at, created_at, updated_at,
    raw_app_meta_data, raw_user_meta_data)
values ('{uid}', '00000000-0000-0000-0000-000000000000', 'authenticated',
    'authenticated', {q(mail(n))}, crypt('kaspio-demo', gen_salt('bf')),
    now(), now(), now(), '{{"provider":"email","providers":["email"]}}'::jsonb,
    {q(json.dumps({"full_name": n}))}::jsonb)
on conflict (id) do nothing;
insert into public.profiles (id, email, full_name)
values ('{uid}', {q(mail(n))}, {q(n)}) on conflict (id) do nothing;""")
A("")

# Eén rij per comité dat iemand trekt, niet één per potje. Een group_owner
# beheert alles in zijn groep, ook de potjes die er later bij komen. Met 118
# posten scheelt dat 118 rijen tegenover 14, en het blijft kloppen als de club
# een post toevoegt.
mrows = []
for name, coms in people.items():
    for com in coms:
        if com not in group_id:
            continue
        mrows.append(f"  ('{ORG}', '{user_id[name]}', 'group_owner', "
                     f"'{group_id[com]}', '{OWNER}')")
A("insert into public.memberships (organisation_id, user_id, role, group_id, invited_by) values")
A(",\n".join(mrows) + "\non conflict do nothing;\n")

open(os.path.join(OUT, "seed_01_structure.sql"), "w").write("\n".join(sql))

# --- transacties
def tx_values(r, pid, tg):
    amt = round(abs(float(r["amount"])), 2)
    direction = "in" if r["amount"] > 0 else "out"
    d = str(r["date"])[:10]
    memo = r["memo"]
    if r["kind"] and r["kind"] != "manueel" and not memo:
        memo = r["descr"]
    cp = r["cp_name"] or (r["kind"] if r["kind"] != "manueel" else None)
    ba = acc(r["own"]) if r["own"] else None
    return (f"('{ORG}', {('null' if pid is None else chr(39)+pid+chr(39))}, "
            f"{amt:.2f}, '{direction}', '{d}', {q(memo)}, {q(cp)}, "
            f"{q(ba)}, {('null' if tg is None else chr(39)+tg+chr(39))})")


# De club begon 2026 niet met een lege rekening. Zonder die openingsstand toont
# het dashboard alleen de beweging van dit jaar, en die is negatief terwijl er
# wel degelijk geld op de rekening stond.
#
# Het bedrag komt NIET uit de cel "EIND DECEMBER 2025 / CASH OP REK", want die
# is in tegenspraak met de rest van het bestand. De maandketen op het blad 2026
# is wél sluitend: elke afgeleide beginstand is de slotstand van de maand
# ervoor. Daaruit volgt 83.460,74 voor 1 januari, en de controle klopt op de
# cent:
#
#     83.460,74 (begin) - 38.923,27 (beweging jan-juli) = 44.537,47
#
# en 44.537,47 is precies wat hun blad als kasstand eind juli noteert. De cel
# zegt 73.098,41; dat is 10.362,33 minder en sluit nergens op aan. Vermoedelijk
# telt die cel niet alle zes rekeningen mee.
OPENING = 83460.74
tx = []
tx.append(f"('{ORG}', '{RESERVE}', {OPENING:.2f}, 'in', '2026-01-01', "
          f"'Kaspositie op 31 december 2025, overgedragen naar 2026', "
          f"'Beginsaldo', null, null)")
for r in recs:
    tx.append(tx_values(r, pot_id[(r["comite"], r["post"])], None))
# Afsluitingsregels zijn geen bankverrichting. Ze horen wel bij hun post, want
# de club rekent ze mee in Stand 2026. Daarom twee benen met hetzelfde
# transfer_group: het bedrag op de post, de tegenboeking op de hoofdpot. De post
# klopt daardoor met hun cijfer, en het totaal van de organisatie blijft gelijk
# aan wat er op de rekening staat. De frontend houdt regels met een
# transfer_group al buiten de in- en uitstroom.
for r in afsl:
    # Handmatige regels hebben geen datum in het blad. Ze horen bij het boekjaar,
    # dus zetten we ze op 31 december; dat is ook waar de club ze afsluit.
    if r["date"] is None:
        r["date"] = datetime.datetime(YEAR, 12, 31)
    g = det("afsluiting", r["comite"], r["post"], str(r["date"])[:10], r["amount"],
            r["memo"] or "")
    amt = round(abs(float(r["amount"])), 2)
    kant = "in" if r["amount"] > 0 else "out"
    tegen = "out" if kant == "in" else "in"
    d = str(r["date"])[:10]
    label = q(r["memo"] or "Afsluiting boekjaar")
    tx.append(f"('{ORG}', '{pot_id[(r['comite'], r['post'])]}', {amt:.2f}, "
              f"'{kant}', '{d}', {label}, 'Interne toewijzing', null, '{g}')")
    tx.append(f"('{ORG}', '{HERVERDEEL}', {amt:.2f}, '{tegen}', '{d}', {label}, "
              f"'Interne toewijzing', null, '{g}')")

for r in losse:
    tx.append(tx_values(r, None, None))
for i, r in enumerate(extra):
    tx.append(tx_values(r, None, tgroup[i]))

HEAD = ("insert into public.transactions (organisation_id, pot_id, amount, "
        "direction, occurred_on, memo, counterparty, bank_account, "
        "transfer_group) values\n")
CHUNK = 250
files = []
for n, i in enumerate(range(0, len(tx), CHUNK)):
    fn = os.path.join(OUT, f"seed_02_tx_{n:03d}.sql")
    open(fn, "w").write(HEAD + ",\n".join(tx[i:i + CHUNK]) + ";\n")
    files.append(fn)

# --- domiciliëringen als recurring_plans
dom = collections.defaultdict(list)
for r in recs:
    if r["kind"] == "Domiciliëring" and r["amount"] < 0 and r["cp_name"]:
        dom[(r["cp_name"].strip(), r["comite"], r["post"])].append(r)
plans = []
for (cp, com, post), rs in dom.items():
    if len(rs) < 3:
        continue
    amt = round(sum(abs(x["amount"]) for x in rs) / len(rs), 2)
    day = collections.Counter(x["date"].day for x in rs).most_common(1)[0][0]
    plans.append(f"  ('{ORG}', '{pot_id[(com, post)]}', 'domiciliering', "
                 f"{amt:.2f}, {day}, {q(cp)}, 5, true)")
p_sql = ""
if plans:
    p_sql = ("insert into public.recurring_plans (organisation_id, pot_id, kind, "
             "amount, day_of_month, counterparty, match_window_days, active) values\n"
             + ",\n".join(plans) + ";\n")
open(os.path.join(OUT, "seed_03_recurring.sql"), "w").write(p_sql)

# =============================================================================
print(f"org            {ORG}")
print(f"pot_groups     {len(GROUP_ORDER)}")
print(f"pots           {len(posts)}")
print(f"leden          {len(user_id)} ({', '.join(user_id)})")
print(f"pot_owner rijen{len(mrows):>4}")
print(f"transacties    {len(tx)}  (gecategoriseerd {len(recs)}, hoofdpot {len(extra)})")
print(f"  interne overboekingen: {npair} paren ({npair*2} regels)")
print(f"  onbeslist in hoofdpot: {len(extra)-npair*2}")
print(f"domiciliëringen{len(plans):>4}")
print(f"tx-bestanden   {len(files)}")
tot_in = sum(r['amount'] for r in recs if r['amount'] > 0)
tot_out = sum(-r['amount'] for r in recs if r['amount'] < 0)
print(f"\ngecategoriseerd 2026: in {tot_in:,.2f}  uit {tot_out:,.2f}  netto {tot_in-tot_out:,.2f}")
json.dump({"org": ORG, "files": files}, open(os.path.join(OUT, "seed_meta.json"), "w"))
EOF_MARKER = None
